"""
compare_to_ground_truth.py

Automates the manual check: given the ground-truth segmentation and the
script's combined output, mark every predicted segment correct or wrong
and colour-highlight it, so no eyeballing is needed.

Usage:
    # two separate files (each uses its first sheet):
    python3 compare_to_ground_truth.py GROUND_TRUTH.xlsx SCRIPT_OUTPUT.xlsx [-o result.xlsx]

    # one combined workbook holding both, chosen by sheet name:
    python3 compare_to_ground_truth.py COMBINED.xlsx \
        --gt-sheet "Ground Truth" --output-sheet Script_Output_v3 [-o result.xlsx]

Matching is by PDF filename (case-insensitive), not the ground-truth's
random PDF ID column. A predicted segment is CORRECT only if its start and
end page exactly match a true document. Everything else is wrong, and the
`mismatch_type` column says how:

    off_by_one      start or end is 1 page off a true document
    over_segmented  a fragment of one true document (an extra split)
    merged          covers 2+ true documents (a missed boundary)
    boundary_shift  overlaps one true document but off by more than a page
    no_ground_truth this PDF wasn't found in the ground-truth sheet
"""
import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font

GREEN = PatternFill("solid", fgColor="C6EFCE")   # correct
RED   = PatternFill("solid", fgColor="FFC7CE")   # wrong
GREY  = PatternFill("solid", fgColor="D9D9D9")   # no ground truth
BOLD  = Font(bold=True)


def norm_name(v):
    """Filename as a stable match key: lowercased, extension dropped."""
    if v is None:
        return None
    s = str(v).strip().lower()
    for ext in (".pdf",):
        if s.endswith(ext):
            s = s[: -len(ext)]
    return s or None


def find_columns(ws):
    """Locate the header row and the columns we need, by header name.

    Both sheets label columns; we don't assume a fixed position. Returns
    (header_row_index, {key: col_index}).
    """
    wanted = {
        "filename": ("pdf_filename", "pdf filename", "filename", "file name"),
        "start":    ("start_page", "start page", "start"),
        "end":      ("end_page", "end page", "end"),
    }
    for r in range(1, min(ws.max_row, 15) + 1):
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            name = str(v).strip().lower()
            for key, aliases in wanted.items():
                if name in aliases and key not in headers:
                    headers[key] = c
        if {"filename", "start", "end"} <= headers.keys():
            return r, headers
    raise ValueError(
        f"Could not find filename/start/end columns in sheet '{ws.title}'. "
        f"Headers seen: {[ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]}"
    )


def pick_sheet(wb, sheet, role):
    """Return the requested sheet, or the first sheet if none named."""
    if sheet is None:
        return wb[wb.sheetnames[0]]
    if sheet not in wb.sheetnames:
        sys.exit(f"{role} sheet '{sheet}' not found. "
                 f"Sheets in this file: {wb.sheetnames}")
    return wb[sheet]


def read_segments(path, sheet=None):
    """{normalised_filename: [(start, end), ...]} from the chosen sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = pick_sheet(wb, sheet, "ground-truth")
    hdr, cols = find_columns(ws)
    out = {}
    for r in range(hdr + 1, ws.max_row + 1):
        fn = norm_name(ws.cell(row=r, column=cols["filename"]).value)
        s = ws.cell(row=r, column=cols["start"]).value
        e = ws.cell(row=r, column=cols["end"]).value
        if fn is None or s is None or e is None:
            continue
        out.setdefault(fn, []).append((int(s), int(e)))
    return out


def classify(seg, truth):
    """Compare one predicted (start, end) against a PDF's true segments.

    Returns (is_correct, mismatch_type). All decisions come from the page
    numbers alone, so there is no guesswork.
    """
    s, e = seg
    true_set = set(truth)
    if (s, e) in true_set:
        return True, "exact"

    # A true document starts strictly inside this segment => a boundary was
    # missed and 2+ documents got merged.
    internal_starts = [gs for (gs, ge) in truth if s < gs <= e]
    if internal_starts:
        return False, "merged"

    # Otherwise this segment sits within (or around) a single true document.
    container = next(((gs, ge) for (gs, ge) in truth if gs <= s and e <= ge), None)
    if container is None:
        container = next(((gs, ge) for (gs, ge) in truth
                          if not (e < gs or s > ge)), None)  # any overlap
    if container is None:
        return False, "boundary_shift"

    gs, ge = container
    if abs(s - gs) <= 1 and abs(e - ge) <= 1:
        return False, "off_by_one"
    if s >= gs and e <= ge:
        return False, "over_segmented"
    return False, "boundary_shift"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("ground_truth", help="ground-truth .xlsx")
    ap.add_argument("script_output", nargs="?",
                    help="script output .xlsx (default: same file as ground_truth, "
                         "for a combined workbook)")
    ap.add_argument("--gt-sheet", help="ground-truth sheet name (default: first sheet)")
    ap.add_argument("--output-sheet", help="script-output sheet name (default: first sheet)")
    ap.add_argument("-o", "--output", default="comparison_result.xlsx")
    args = ap.parse_args()

    # One combined workbook is the common case: pass it once, pick sheets by name.
    script_path = args.script_output or args.ground_truth
    for p in {args.ground_truth, script_path}:
        if not Path(p).exists():
            sys.exit(f"File not found: {p}")

    truth = read_segments(args.ground_truth, args.gt_sheet)
    print(f"Ground truth: {len(truth)} PDFs, "
          f"{sum(len(v) for v in truth.values())} true segments")

    wb = openpyxl.load_workbook(script_path)   # keep styles/formatting
    ws = pick_sheet(wb, args.output_sheet, "script-output")
    hdr, cols = find_columns(ws)

    # Append two result columns.
    res_col = ws.max_column + 1
    why_col = res_col + 1
    ws.cell(row=hdr, column=res_col, value="check_result").font = BOLD
    ws.cell(row=hdr, column=why_col, value="mismatch_type").font = BOLD

    counts = {"correct": 0, "wrong": 0, "no_ground_truth": 0}
    per_pdf = {}
    matched_pdfs = set()

    for r in range(hdr + 1, ws.max_row + 1):
        fn = norm_name(ws.cell(row=r, column=cols["filename"]).value)
        s = ws.cell(row=r, column=cols["start"]).value
        e = ws.cell(row=r, column=cols["end"]).value
        if fn is None or s is None or e is None:
            continue

        if fn not in truth:
            fill, result, why = GREY, "no_ground_truth", "pdf_not_in_ground_truth"
            counts["no_ground_truth"] += 1
        else:
            matched_pdfs.add(fn)
            ok, why = classify((int(s), int(e)), truth[fn])
            if ok:
                fill, result = GREEN, "correct"
                counts["correct"] += 1
            else:
                fill, result = RED, "wrong"
                counts["wrong"] += 1
            d = per_pdf.setdefault(fn, {"correct": 0, "wrong": 0})
            d["correct" if ok else "wrong"] += 1

        ws.cell(row=r, column=res_col, value=result)
        ws.cell(row=r, column=why_col, value=why)
        for c in range(1, why_col + 1):
            ws.cell(row=r, column=c).fill = fill

    # A second sheet: per-PDF scorecard. Replace any stale one from a re-run.
    if "comparison_summary" in wb.sheetnames:
        del wb["comparison_summary"]
    sm = wb.create_sheet("comparison_summary")
    sm.append(["pdf_filename", "predicted", "correct", "wrong", "accuracy_%"])
    for c in sm[1]:
        c.font = BOLD
    for fn in sorted(per_pdf):
        d = per_pdf[fn]
        n = d["correct"] + d["wrong"]
        sm.append([fn, n, d["correct"], d["wrong"],
                   round(d["correct"] / n * 100, 1) if n else 0])
    tot = counts["correct"] + counts["wrong"]
    sm.append([])
    row = ["TOTAL", tot, counts["correct"], counts["wrong"],
           round(counts["correct"] / tot * 100, 1) if tot else 0]
    sm.append(row)
    for c in sm[sm.max_row]:
        c.font = BOLD

    # Ground-truth PDFs the script produced no output for. These never
    # appear as red rows (there is no output row to colour), so they would
    # be a silent gap. List them explicitly.
    unmatched = sorted(set(truth) - matched_pdfs)
    if unmatched:
        sm.append(["", "", "", "", ""])   # spacer (non-empty so max_row advances)
        sm.append(["Ground-truth PDFs with NO script output (missing entirely):"])
        hdr_row = sm.max_row
        sm.cell(row=hdr_row, column=1).font = BOLD
        sm.cell(row=hdr_row, column=1).fill = GREY
        for fn in unmatched:
            sm.append([fn, "no output produced", len(truth[fn]), "", ""])
            sm.cell(row=sm.max_row, column=1).fill = GREY

    wb.save(args.output)

    print(f"\nPredicted segments checked: {tot}")
    print(f"  correct : {counts['correct']} ({counts['correct']/tot*100:.1f}%)")
    print(f"  wrong   : {counts['wrong']}")
    if counts["no_ground_truth"]:
        print(f"  no ground truth for: {counts['no_ground_truth']} segments")
    if unmatched:
        print(f"  ground-truth PDFs with no script output: {len(unmatched)}")
    print(f"\nWrote {args.output}  (highlighted rows + a comparison_summary sheet)")


if __name__ == "__main__":
    main()
